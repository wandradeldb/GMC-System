"""
Import historical PayApp data from:
  G&D 17_204_001_163_ PayApp_MerlinPk (#2) April 26_ER Determination.xlsx

PayApp #1 — reconstructed from PayApp #2 'Previously Certified' figure
PayApp #2 — full detail from Application + Application Detail sheets
"""
import sqlite3, os
from openpyxl import load_workbook
from datetime import datetime

XLSX = r"C:\Users\wagne\Downloads\G&D 17_204_001_163_ PayApp_MerlinPk (#2) April 26_ER Determination.xlsx"
DB   = os.path.join(os.path.dirname(__file__), "gmc.db")
PROJECT_ID = 1

# Apply schema migration first
print("Applying migration 004_payapp.sql...")
migration = open(os.path.join(os.path.dirname(__file__), "migrations", "004_payapp.sql")).read()
con = sqlite3.connect(DB)
con.execute("PRAGMA foreign_keys = ON")
# Split on CREATE/INSERT statements safely
import re
statements = [s.strip() for s in re.split(r';\s*\n', migration) if s.strip()]
for stmt in statements:
    try:
        con.execute(stmt)
    except Exception as e:
        if 'already exists' not in str(e):
            print(f"  Warning: {e}")
con.commit()
print("  Done.")

# ── Read Excel ─────────────────────────────────────────────────────────────
print(f"\nReading {XLSX}...")
wb = load_workbook(XLSX, data_only=True)
ws_app = wb['Application']

def cell_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return float(v) if v is not None and isinstance(v, (int, float)) else 0.0

# PayApp #2 values from Application sheet
pa2_date_submitted      = ws_app.cell(4, 11).value   # datetime or string
pa2_works_gross_cum     = cell_val(ws_app, 24, 10)   # Works Present Valuation = 297,048.57
pa2_ae_cum              = cell_val(ws_app, 25, 10)   # Adjustment Events = 0
pa2_total_gross_cum     = cell_val(ws_app, 30, 10)   # Total = 297,048.57
pa2_retention_pct       = cell_val(ws_app, 31, 7) * 100  # 0.03 → 3.0
pa2_retention_cum       = cell_val(ws_app, 31, 10)   # = 8,911.46
pa2_net_cumulative      = cell_val(ws_app, 32, 10)   # = 288,137.12
pa2_previously_cert     = cell_val(ws_app, 33, 10)   # = 77,766.73
pa2_this_certificate    = cell_val(ws_app, 35, 10)   # = 210,370.39

if isinstance(pa2_date_submitted, datetime):
    pa2_date_str = pa2_date_submitted.strftime('%Y-%m-%d')
else:
    pa2_date_str = '2026-05-05'

print(f"\nPayApp #2 values from Excel:")
print(f"  Date submitted       : {pa2_date_str}")
print(f"  Works gross cum      : €{pa2_works_gross_cum:,.2f}")
print(f"  Retention @{pa2_retention_pct:.1f}%      : €{pa2_retention_cum:,.2f}")
print(f"  Net cumulative       : €{pa2_net_cumulative:,.2f}")
print(f"  Previously certified : €{pa2_previously_cert:,.2f}")
print(f"  This certificate     : €{pa2_this_certificate:,.2f}")

# ── Reconstruct PayApp #1 from PayApp #2's 'Previously Certified' ──────────
# PayApp #1 net = previously_certified = 77,766.73
# If retention_pct = 3%, gross_pa1 = net / 0.97 = 80,172.92
pa1_net             = pa2_previously_cert
pa1_gross           = round(pa1_net / (1 - pa2_retention_pct / 100), 2)
pa1_retention       = round(pa1_gross * pa2_retention_pct / 100, 2)

print(f"\nPayApp #1 (reconstructed):")
print(f"  Works gross cum      : €{pa1_gross:,.2f}")
print(f"  Retention            : €{pa1_retention:,.2f}")
print(f"  Net cumulative       : €{pa1_net:,.2f}")
print(f"  This certificate     : €{pa1_net:,.2f}")

# ── Read item-level detail for PayApp #2 ───────────────────────────────────
ws_det = wb['Application Detail']
item_rows = []
for row in ws_det.iter_rows():
    col5 = row[4].value   # Type (E)
    col2 = row[1].value   # Item ref (B)
    col4 = row[3].value   # Contract sum (D)
    col10 = row[9].value  # PayApp #1 pct (J)
    col11 = row[10].value # PayApp #2 pct (K)

    if col5 in ('F','T','M') and col4 and isinstance(col4, (int, float)):
        pct_pa1 = float(col10) * 100 if col10 is not None and isinstance(col10,(int,float)) else None
        pct_pa2 = float(col11) * 100 if col11 is not None and isinstance(col11,(int,float)) else None
        item_rows.append({'ref': str(col2) if col2 else '', 'cs': float(col4), 'pct_pa1': pct_pa1, 'pct_pa2': pct_pa2})

# Look up boq_item IDs by item_ref
cur = con.execute("SELECT id, item_ref FROM boq_item WHERE project_id=?", (PROJECT_ID,))
ref_to_id = {row[1]: row[0] for row in cur.fetchall()}

# ── Upsert PayApp #1 ───────────────────────────────────────────────────────
con.execute("BEGIN")
con.execute("""
    INSERT INTO payapp (project_id, app_number, period, date_submitted, status,
        works_gross_cumulative, ae_cumulative, total_gross_cumulative,
        retention_pct, retention_cumulative, net_cumulative,
        previously_certified, this_certificate, source, notes)
    VALUES (?,1,'2026-02','2026-02-01','certified',?,0,?,?,?,?,0,?,
        'import','Reconstructed from PayApp #2 Previously Certified figure')
    ON CONFLICT(project_id,app_number) DO UPDATE SET
        works_gross_cumulative=excluded.works_gross_cumulative,
        retention_cumulative=excluded.retention_cumulative,
        net_cumulative=excluded.net_cumulative,
        this_certificate=excluded.this_certificate,
        status=excluded.status
""", (PROJECT_ID, pa1_gross, pa1_gross, pa2_retention_pct, pa1_retention, pa1_net, pa1_net))

pa1_id = con.execute("SELECT id FROM payapp WHERE project_id=? AND app_number=1", (PROJECT_ID,)).fetchone()[0]
print(f"\nInserted PayApp #1 (id={pa1_id})")

# ── Upsert PayApp #2 ───────────────────────────────────────────────────────
con.execute("""
    INSERT INTO payapp (project_id, app_number, period, date_submitted, status,
        works_gross_cumulative, ae_cumulative, total_gross_cumulative,
        retention_pct, retention_cumulative, net_cumulative,
        previously_certified, this_certificate, source, notes)
    VALUES (?,2,'2026-05',?,
        CASE WHEN ? > 0 THEN 'certified' ELSE 'submitted' END,
        ?,0,?,?,?,?,?,?,
        'import','Imported from G&D PayApp #2 April 2026 ER Determination')
    ON CONFLICT(project_id,app_number) DO UPDATE SET
        works_gross_cumulative=excluded.works_gross_cumulative,
        retention_cumulative=excluded.retention_cumulative,
        net_cumulative=excluded.net_cumulative,
        previously_certified=excluded.previously_certified,
        this_certificate=excluded.this_certificate,
        status=excluded.status
""", (PROJECT_ID, pa2_date_str,
      pa2_this_certificate,
      pa2_works_gross_cum, pa2_total_gross_cum,
      pa2_retention_pct, pa2_retention_cum, pa2_net_cumulative,
      pa2_previously_cert, pa2_this_certificate))

pa2_id = con.execute("SELECT id FROM payapp WHERE project_id=? AND app_number=2", (PROJECT_ID,)).fetchone()[0]
print(f"Inserted PayApp #2 (id={pa2_id})")

# ── Insert item detail for PayApps ─────────────────────────────────────────
items_inserted = 0
for item in item_rows:
    bid = ref_to_id.get(item['ref'])
    if not bid:
        continue

    # PayApp #1 item
    if item['pct_pa1'] is not None and item['pct_pa1'] > 0:
        val = round(item['cs'] * item['pct_pa1'] / 100, 4)
        con.execute("""
            INSERT INTO payapp_item (payapp_id, boq_item_id, pct_complete, value_claimed)
            VALUES (?,?,?,?)
            ON CONFLICT(payapp_id, boq_item_id) DO UPDATE SET pct_complete=excluded.pct_complete, value_claimed=excluded.value_claimed
        """, (pa1_id, bid, item['pct_pa1'], val))
        items_inserted += 1

    # PayApp #2 item — use pa2 if set, else inherit pa1
    pct_pa2 = item['pct_pa2'] if item['pct_pa2'] is not None else item['pct_pa1']
    if pct_pa2 is not None and pct_pa2 >= 0:
        val = round(item['cs'] * pct_pa2 / 100, 4)
        con.execute("""
            INSERT INTO payapp_item (payapp_id, boq_item_id, pct_complete, value_claimed)
            VALUES (?,?,?,?)
            ON CONFLICT(payapp_id, boq_item_id) DO UPDATE SET pct_complete=excluded.pct_complete, value_claimed=excluded.value_claimed
        """, (pa2_id, bid, pct_pa2, val))
        items_inserted += 1

con.execute("COMMIT")
con.close()

print(f"Inserted {items_inserted} payapp_item records")
print("\nDone. Historical PayApp data imported.")
print(f"\nSummary:")
print(f"  PayApp #1 (Feb 2026): gross €{pa1_gross:,.2f}  →  net cert €{pa1_net:,.2f}")
print(f"  PayApp #2 (May 2026): gross €{pa2_works_gross_cum:,.2f}  →  net cert €{pa2_this_certificate:,.2f}")
print(f"  Total certified to date: €{pa2_net_cumulative:,.2f}")
